from importlib.util import module_from_spec, spec_from_file_location
from datetime import date
from pathlib import Path
import sys

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
sys.path.insert(0, str(SCRIPTS))
SPEC = spec_from_file_location("import_legacy_costs", SCRIPTS / "import_legacy_costs.py")
MODULE = module_from_spec(SPEC)
sys.modules[SPEC.name] = MODULE
SPEC.loader.exec_module(MODULE)


def test_cost_import_excludes_la_nave_and_mixed_rows(tmp_path):
    book = Workbook()
    baiamonte = book.active
    baiamonte.title = "Gennaio 2024"
    baiamonte.append(["DATA", "OGGETTO", "IMPORTO ", "TIPO FATTURA"])
    baiamonte.append([None, "Baiamonte poles", 100, "TNB"])
    baiamonte.append([None, "Mixed fuel", 50, "TNB - LA NAVE"])
    la_nave = book.create_sheet("s.s. LA NAVE 2023")
    la_nave.append(["La Nave expense", 999, "2023-01-01"])
    path = tmp_path / "Baiamonte Costs Worksheet.xlsx"
    book.save(path)

    records = MODULE.parse_cost_book(path)

    assert [row.description for row in records] == ["Baiamonte poles"]
    assert records[0].amount_eur == 100
    assert records[0].record_date.isoformat() == "2024-01-01"


def test_decimal_comma_and_ture_month_duplicate_are_reconciled(tmp_path):
    book = Workbook()
    expenses = book.active
    expenses.title = "spese sistemazione baia monte "
    expenses.append(["sistemazione Baia monte", "costo", "data", "pagamento"])
    expenses.append(["TURE - lavoro vigna mese luglio 2023", "530,31", "04-08-23", "contanti"])
    ture = book.create_sheet("Ture-2023")
    for _ in range(4):
        ture.append([])
    ture.append([None, "Ture Bill", 0, 0, 0, 0, 0, 0, 530.31])
    path = tmp_path / "Baiamonte Costs Worksheet.xlsx"
    book.save(path)

    records = MODULE.parse_cost_book(path)
    MODULE.reconcile(records)

    assert records[0].amount_eur == 530.31
    duplicate = next(row for row in records if row.source_sheet == "Ture-2023")
    assert duplicate.included_in_totals is False
    assert duplicate.duplicate_of == records[0].id


def test_history_api_includes_baiamonte_expenses():
    source = (ROOT / "app" / "main.py").read_text()
    index = (ROOT / "app" / "static" / "index.html").read_text()
    assert '"historical_costs": "SELECT record_year year' in source
    assert "/api/v1/admin/import-workbooks" not in source
    assert "workbookImportForm" not in index


def test_italian_slash_dates_and_month_precision_are_preserved():
    assert MODULE.parse_date("23/8/25") == (date(2025, 8, 23), "day")
    assert MODULE.parse_date("8/9/2025") == (date(2025, 9, 8), "day")
    assert MODULE.parse_date("06/23/25") == (date(2025, 6, 23), "day")
    assert MODULE.parse_date("12/2024") == (date(2024, 12, 1), "month")


def test_explicit_legacy_hours_are_retained_as_labor_evidence():
    record = MODULE.Record("source", "work.xlsx", "PAYMENTS", 20, "Giancarlo 47 ore novembre", 470)
    assert record.labor_hours == 47


def test_existing_legacy_dates_and_hours_are_backfilled_by_migrations():
    dates = (ROOT / "db" / "migrations" / "042_repair_legacy_work_dates.sql").read_text()
    hours = (ROOT / "db" / "migrations" / "043_historical_labor_hours.sql").read_text()
    assert "STR_TO_DATE(raw_date,'%d/%m/%y')" in dates
    assert "ADD COLUMN IF NOT EXISTS labor_hours" in hours


def test_proclama_payroll_adds_labor_without_duplicating_costs_or_payment_claims():
    migration = (ROOT / "db" / "migrations" / "049_giancarlo_proclama_labor.sql").read_text()
    assert "Proclama payroll archive" in migration
    assert "0,45.50,'payment_not_verified',0" in migration
    assert "0,123.50,'payment_not_verified',0" in migration
    assert "'2025-11-30',2025,'month'" in migration
    assert "15 worked days, hours not stated" in migration
    assert "0,NULL,'payment_not_verified',0" in migration
    assert migration.count(";") == 1
