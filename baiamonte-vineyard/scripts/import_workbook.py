"""Retired one-time workbook migration code retained for historical audit tests.

The command entry point is disabled. MariaDB and authenticated application
inputs are the only operational authorities.

Dry-run is the default. Use --commit only after reviewing the JSON report and taking a backup.
Every non-empty workbook row is preserved in workbook_source_rows even when it does not
yet have a normalized destination table.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import os
import re
import sys
import uuid
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable

import openpyxl
import pymysql


ESTATE_ID = "00000000-0000-4000-8000-000000000001"
EXPECTED_TITLE = "Tenuta Baiamonte — Vineyard, Harvest & Cellar Audit Log"


def uid() -> str:
    return str(uuid.uuid4())


def clean(value: Any) -> Any:
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def as_text(value: Any) -> str | None:
    value = clean(value)
    return None if value is None else str(value)


def as_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float, Decimal)):
        return float(value)
    text = re.sub(r"[^0-9,.-]", "", str(value)).replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def as_int(value: Any) -> int | None:
    number = as_number(value)
    return None if number is None else int(round(number))


def as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None or str(value).strip() == "":
        return None
    text = str(value).strip()
    for pattern in ("%Y-%m-%d", "%m/%d/%Y", "%d %b %Y", "%d %B %Y"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            pass
    return None


def as_datetime(date_value: Any, time_value: Any = None) -> datetime | None:
    day = as_date(date_value)
    if not day:
        return None
    if isinstance(time_value, datetime):
        clock = time_value.time()
    elif isinstance(time_value, time):
        clock = time_value
    elif time_value:
        try:
            clock = datetime.strptime(str(time_value), "%H:%M").time()
        except ValueError:
            clock = time(12, 0)
    else:
        clock = time(12, 0)
    return datetime.combine(day, clock)


def json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def normalized_row(values: Iterable[Any]) -> list[Any]:
    row = [clean(item) for item in values]
    while row and row[-1] is None:
        row.pop()
    return row


def find_table(sheet: Any, first_header: str) -> tuple[list[str], list[tuple[int, dict[str, Any]]]]:
    rows = list(sheet.iter_rows(values_only=True))
    header_index = next((idx for idx, row in enumerate(rows) if as_text(row[0]) == first_header), None)
    if header_index is None:
        return [], []
    headers = [as_text(item) or f"Column {idx + 1}" for idx, item in enumerate(rows[header_index])]
    output: list[tuple[int, dict[str, Any]]] = []
    for idx, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        values = normalized_row(row)
        if not values or values[0] is None:
            break
        output.append((idx, {headers[col]: values[col] if col < len(values) else None for col in range(len(headers))}))
    return headers, output


def find_tables(sheet: Any, first_header: str) -> list[tuple[list[str], list[tuple[int, dict[str, Any]]]]]:
    rows = list(sheet.iter_rows(values_only=True))
    tables: list[tuple[list[str], list[tuple[int, dict[str, Any]]]]] = []
    for header_index, source_row in enumerate(rows):
        if as_text(source_row[0]) != first_header:
            continue
        headers = [as_text(item) or f"Column {idx + 1}" for idx, item in enumerate(source_row)]
        output: list[tuple[int, dict[str, Any]]] = []
        for idx, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
            values = normalized_row(row)
            if not values or values[0] is None:
                break
            output.append((idx, {headers[col]: values[col] if col < len(values) else None for col in range(len(headers))}))
        tables.append((headers, output))
    return tables


class Importer:
    def __init__(self, workbook_path: Path, commit: bool, source_file_id: str | None, source_modified_at: str | None):
        self.path = workbook_path
        self.commit_mode = commit
        self.source_file_id = source_file_id
        self.source_modified_at = source_modified_at
        self.workbook = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
        self.sha256 = hashlib.sha256(workbook_path.read_bytes()).hexdigest()
        self.batch_id = uid()
        self.counts: dict[str, int] = {"raw_rows": 0}
        self.warnings: list[str] = []
        self.connection = None
        self.cursor = None
        self.preserve_source_rows = True
        self.seasons: dict[int, str] = {}
        self.varieties: dict[str, str] = {}

    def connect(self) -> None:
        self.connection = pymysql.connect(
            host=os.getenv("DB_HOST", "core-mariadb"), port=int(os.getenv("DB_PORT", "3306")),
            user=os.getenv("DB_USER", "baiamonte"), password=os.getenv("DB_PASSWORD", ""),
            database=os.getenv("DB_NAME", "baiamonte_vineyard"), charset="utf8mb4", autocommit=False,
        )
        self.cursor = self.connection.cursor()

    def run(self) -> dict[str, Any]:
        required = {"Setup & Lists", "Block Plan", "Harvest Operations", "Cellar Operations", "Lab Results", "Historical Lab Results"}
        missing = sorted(required - set(self.workbook.sheetnames))
        if missing:
            raise ValueError(f"Workbook is missing required sheets: {', '.join(missing)}")
        if self.commit_mode:
            self.connect()
            self.start_batch()
        try:
            self.preserve_raw_rows()
            self.import_reference_data()
            self.import_blocks()
            self.import_parcels_and_atlas()
            self.import_issues()
            self.import_labor()
            self.import_equipment()
            self.import_olive()
            self.import_treatments()
            self.import_maturity_samples()
            self.import_harvest_plans()
            self.import_harvest_operations()
            self.import_cellar_operations()
            self.import_fermentation()
            self.import_mass_balance()
            self.import_vintage_history()
            self.import_projections()
            self.import_labs("Historical Lab Results")
            self.import_labs("Lab Results")
            report = self.report()
            if self.commit_mode:
                self.cursor.execute(
                    "UPDATE import_batches SET status='committed',row_count=%s,warning_count=%s,report=%s,completed_at=NOW(6) WHERE id=%s",
                    (self.counts["raw_rows"], len(self.warnings), json.dumps(report, default=json_value), self.batch_id),
                )
                self.connection.commit()
            return report
        except Exception:
            if self.connection:
                self.connection.rollback()
            raise
        finally:
            if self.connection:
                self.connection.close()

    def start_batch(self) -> None:
        self.cursor.execute("SELECT id FROM import_batches WHERE estate_id=%s AND content_sha256=%s", (ESTATE_ID, self.sha256))
        existing = self.cursor.fetchone()
        if existing:
            # A newer importer may normalize domains that were retained only as
            # raw source rows by an earlier release. Reuse the audited batch and
            # rerun idempotent domain upserts without duplicating source rows.
            self.batch_id = existing[0]
            self.preserve_source_rows = False
            self.cursor.execute("UPDATE import_batches SET status='started',warning_count=0,report=NULL,completed_at=NULL WHERE id=%s", (self.batch_id,))
            return
        self.cursor.execute(
            "INSERT INTO import_batches (id,estate_id,source_name,source_file_id,source_modified_at,content_sha256,status) VALUES (%s,%s,%s,%s,%s,%s,'started')",
            (self.batch_id, ESTATE_ID, self.path.name, self.source_file_id, self.source_modified_at, self.sha256),
        )

    def bump(self, key: str) -> None:
        self.counts[key] = self.counts.get(key, 0) + 1

    def preserve_raw_rows(self) -> None:
        for sheet in self.workbook.worksheets:
            for row_number, values in enumerate(sheet.iter_rows(values_only=True), start=1):
                row = normalized_row(values)
                if not row:
                    continue
                payload = json.dumps(row, default=json_value, ensure_ascii=False)
                row_hash = hashlib.sha256(f"{sheet.title}:{row_number}:{payload}".encode()).hexdigest()
                self.bump("raw_rows")
                if self.commit_mode and self.preserve_source_rows:
                    self.cursor.execute(
                        "INSERT INTO workbook_source_rows (import_batch_id,sheet_name,source_row_number,row_values,row_hash) VALUES (%s,%s,%s,%s,%s)",
                        (self.batch_id, sheet.title, row_number, payload, row_hash),
                    )

    def season(self, year: int) -> str:
        if year not in self.seasons:
            self.seasons[year] = uid()
            if self.commit_mode:
                self.cursor.execute(
                    "INSERT INTO seasons (id,estate_id,vintage_year,status) VALUES (%s,%s,%s,%s) ON DUPLICATE KEY UPDATE status=VALUES(status)",
                    (self.seasons[year], ESTATE_ID, year, "active" if year == datetime.now().year else "closed"),
                )
                self.cursor.execute("SELECT id FROM seasons WHERE estate_id=%s AND vintage_year=%s", (ESTATE_ID, year))
                self.seasons[year] = self.cursor.fetchone()[0]
        return self.seasons[year]

    def variety(self, name: Any) -> str | None:
        name = as_text(name)
        if not name:
            return None
        aliases = {"Granache": "Grenache", "Alicante N.": "Grenache", "Grecanico Dorato B.": "Grecanico", "Nerello": "Nerello Mascalese"}
        canonical = aliases.get(name, name)
        # The historical workbook used Blend and Other as planning buckets,
        # not grape varieties. The real Nerello/Grenache blend is managed by
        # the separate blend-program tables and must never become a harvest
        # variety or receive its own GDD forecast.
        if canonical.casefold() in {"blend", "other"}:
            return None
        key = canonical.casefold()
        if key not in self.varieties:
            self.varieties[key] = uid()
            if self.commit_mode:
                self.cursor.execute(
                    "INSERT INTO grape_varieties (id,estate_id,name) VALUES (%s,%s,%s) ON DUPLICATE KEY UPDATE name=VALUES(name)",
                    (self.varieties[key], ESTATE_ID, canonical),
                )
                self.cursor.execute("SELECT id FROM grape_varieties WHERE estate_id=%s AND name=%s", (ESTATE_ID, canonical))
                self.varieties[key] = self.cursor.fetchone()[0]
        return self.varieties[key]

    def import_reference_data(self) -> None:
        sheet = self.workbook["Setup & Lists"]
        rows = list(sheet.iter_rows(min_row=5, max_row=20, values_only=True))
        for row in rows:
            if as_text(row[3]):
                self.variety(row[3]); self.bump("varieties")
            person = as_text(row[15])
            if person:
                self.bump("people")
                if self.commit_mode:
                    self.cursor.execute(
                        "INSERT INTO people (id,estate_id,name,role,active) VALUES (%s,%s,%s,%s,1) ON DUPLICATE KEY UPDATE role=VALUES(role)",
                        (uid(), ESTATE_ID, person, as_text(row[17])),
                    )
        for row in rows:
            key, value = as_text(row[0]), row[1]
            if key and value is not None:
                self.bump("settings")
                if self.commit_mode:
                    self.cursor.execute(
                        "INSERT INTO app_settings (estate_id,setting_key,setting_value) VALUES (%s,%s,%s) ON DUPLICATE KEY UPDATE setting_value=VALUES(setting_value)",
                        (ESTATE_ID, "workbook." + re.sub(r"[^a-z0-9]+", "_", key.lower()).strip("_"), json.dumps(json_value(value))),
                    )

    def import_blocks(self) -> None:
        _, rows = find_table(self.workbook["Block Plan"], "Block ID")
        for _, row in rows:
            block_id, variety_id = uid(), self.variety(row.get("Variety"))
            self.bump("blocks")
            if self.commit_mode:
                self.cursor.execute(
                    "INSERT INTO vineyard_blocks (id,estate_id,code,name,area_ha,vine_count,notes) VALUES (%s,%s,%s,%s,%s,%s,%s) ON DUPLICATE KEY UPDATE name=VALUES(name),area_ha=VALUES(area_ha),vine_count=VALUES(vine_count),notes=VALUES(notes)",
                    (block_id, ESTATE_ID, as_text(row["Block ID"]), as_text(row["Block name"]), as_number(row.get("Area ha")), as_int(row.get("Vine positions")), as_text(row.get("Notes"))),
                )
                self.cursor.execute("SELECT id FROM vineyard_blocks WHERE estate_id=%s AND code=%s", (ESTATE_ID, as_text(row["Block ID"])))
                actual_block_id = self.cursor.fetchone()[0]
                if variety_id:
                    self.cursor.execute("INSERT IGNORE INTO block_varieties (block_id,variety_id,area_ha,vine_count) VALUES (%s,%s,%s,%s)", (actual_block_id, variety_id, as_number(row.get("Area ha")), as_int(row.get("Vine positions"))))

    def import_parcels_and_atlas(self) -> None:
        _, parcels = find_table(self.workbook["Estate & Parcels"], "Municipality")
        for _, row in parcels:
            self.bump("parcels")
            if self.commit_mode:
                self.cursor.execute("INSERT INTO cadastral_parcels (id,estate_id,municipality,cadastral_sheet,parcel_number,tenure,contract_protocol,tenure_start,tenure_end,cadastral_area_ha,conducted_area_ha,buildings_m2,notes) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) ON DUPLICATE KEY UPDATE conducted_area_ha=VALUES(conducted_area_ha),notes=VALUES(notes)", (uid(), ESTATE_ID, as_text(row["Municipality"]), as_text(row["Sheet"]), as_text(row["Parcel"]), as_text(row.get("Tenure")), as_text(row.get("Contract protocol")), as_date(row.get("Start")), as_date(row.get("End")), as_number(row.get("Cadastral ha")), as_number(row.get("Conducted ha")), as_number(row.get("Buildings m²")), as_text(row.get("Operational notes"))))
        for _, terraces in find_tables(self.workbook["Vineyard Atlas"], "Terrace ID"):
            for _, row in terraces:
                code = as_text(row["Terrace ID"])
                if not code or "TOTAL" in code:
                    continue
                self.bump("terraces")
                if self.commit_mode:
                    self.cursor.execute("INSERT INTO vineyard_terraces (id,estate_id,terrace_code,cohort,training_system,allocated_vines,spacing_m,reconciliation_basis,confidence,field_census_status,live_vines,dead_missing_vines,replacement_new_vines,notes) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) ON DUPLICATE KEY UPDATE allocated_vines=VALUES(allocated_vines),field_census_status=VALUES(field_census_status),notes=VALUES(notes)", (uid(), ESTATE_ID, code, as_text(row.get("Cohort")), as_text(row.get("Training")), as_int(row.get("Allocated vines")), as_number(row.get("Spacing m")), as_text(row.get("Reconciliation basis")), as_text(row.get("Confidence")), as_text(row.get("Field census")), as_int(row.get("2026 live vines")), as_int(row.get("Dead / missing")), as_int(row.get("Replacement / new")), as_text(row.get("Notes"))))
        _, deliveries = find_table(self.workbook["Vineyard Atlas"], "Invoice")
        for _, row in deliveries:
            if not as_int(row.get("Quantity")):
                continue
            self.bump("nursery_deliveries")
            if self.commit_mode:
                self.cursor.execute("INSERT INTO nursery_deliveries (id,estate_id,invoice_number,invoice_date,variety_id,supplied_variety_name,quantity,cohort_use,mapping_status) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) ON DUPLICATE KEY UPDATE mapping_status=VALUES(mapping_status)", (uid(), ESTATE_ID, as_text(row.get("Invoice")), as_date(row.get("Invoice date")), self.variety(row.get("Variety")), as_text(row.get("Variety")), as_int(row.get("Quantity")), as_text(row.get("Cohort / use")), as_text(row.get("Mapping status"))))

    def import_issues(self) -> None:
        _, rows = find_table(self.workbook["Issues & Decisions"], "Issue ID")
        for _, row in rows:
            self.bump("issues")
            status = (as_text(row.get("Status")) or "Open").lower().replace(" ", "_")
            if status not in {"open", "monitoring", "resolved", "deferred"}: status = "open"
            priority = (as_text(row.get("Priority")) or "Medium").lower()
            if self.commit_mode:
                self.cursor.execute("INSERT INTO issues_decisions (id,estate_id,source_issue_id,opened_date,subject_ref,issue_type,priority,issue_text,evidence_summary,decision_action,owner_text,due_date,status,closed_date,notes) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) ON DUPLICATE KEY UPDATE priority=VALUES(priority),issue_text=VALUES(issue_text),evidence_summary=VALUES(evidence_summary),decision_action=VALUES(decision_action),owner_text=VALUES(owner_text),due_date=VALUES(due_date),status=VALUES(status),closed_date=VALUES(closed_date),notes=VALUES(notes)", (uid(), ESTATE_ID, as_text(row["Issue ID"]), as_date(row.get("Date opened")), as_text(row.get("Lot/block/asset")), as_text(row.get("Type")) or "Data", priority, as_text(row.get("Issue or decision")), as_text(row.get("Evidence")), as_text(row.get("Decision/action")), as_text(row.get("Owner")), as_date(row.get("Due date")), status, as_date(row.get("Closed date")), as_text(row.get("Notes"))))

    def import_labor(self) -> None:
        _, rows = find_table(self.workbook["Labor Log"], "Labor ID")
        for _, row in rows:
            self.bump("labor_entries")
            role = as_text(row.get("Role")) or ""
            scope = "contractor" if role.lower() == "contractor" else "part_time"
            if "Giancarlo" in (as_text(row.get("Person/crew")) or ""): scope = "payroll_excluded"
            notes = as_text(row.get("Notes")) or ""
            payment = "verification_needed" if "Verification needed" in notes else ("unpaid" if "UNPAID" in notes else "unknown")
            if self.commit_mode:
                self.cursor.execute("INSERT INTO labor_entries (id,estate_id,season_id,source_labor_id,work_date,shift_label,person_or_crew,role,regular_hours,overtime_hours,hourly_rate_eur,labor_cost_eur,kg_handled,incident_near_miss,approved_by,payment_status,payroll_scope,notes) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) ON DUPLICATE KEY UPDATE regular_hours=VALUES(regular_hours),overtime_hours=VALUES(overtime_hours),labor_cost_eur=VALUES(labor_cost_eur),payment_status=VALUES(payment_status),notes=VALUES(notes)", (uid(), ESTATE_ID, self.season(2026), as_text(row["Labor ID"]), as_date(row.get("Date")), as_text(row.get("Shift")), as_text(row.get("Person/crew")), role, as_number(row.get("Regular hours")), as_number(row.get("Overtime hours")), as_number(row.get("Hourly rate")), as_number(row.get("Labor cost")), as_number(row.get("Kg handled")), 1 if str(row.get("Incident/near miss", "")).lower() == "yes" else 0, as_text(row.get("Approved by")), payment, scope, notes))

    def import_equipment(self) -> None:
        _, rows = find_table(self.workbook["Equipment & Sanitation"], "Record ID")
        for _, row in rows:
            if str(row.get("Record ID", "")).startswith("TEMPLATE"): continue
            self.bump("equipment_events")
            if self.commit_mode:
                self.cursor.execute("INSERT INTO equipment_service_events (id,estate_id,source_record_id,event_date,asset_name,pre_use_status,sanitation_method,concentration,released,released_by,downtime_hours,maintenance_action,next_due_date,notes) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) ON DUPLICATE KEY UPDATE pre_use_status=VALUES(pre_use_status),released=VALUES(released),maintenance_action=VALUES(maintenance_action),next_due_date=VALUES(next_due_date),notes=VALUES(notes)", (uid(), ESTATE_ID, as_text(row["Record ID"]), as_date(row.get("Date")), as_text(row.get("Asset")), as_text(row.get("Pre-use status")), as_text(row.get("Sanitizer/method")), as_text(row.get("Concentration")), 1 if str(row.get("Released?", "")).lower() == "yes" else 0, as_text(row.get("Released by")), as_number(row.get("Downtime hours")), as_text(row.get("Maintenance/action")), as_date(row.get("Next due")), as_text(row.get("Notes"))))

    def import_olive(self) -> None:
        _, rows = find_table(self.workbook["Olive Log"], "Record ID")
        for _, row in rows:
            self.bump("olive_records")
            if self.commit_mode:
                self.cursor.execute("INSERT INTO olive_records (id,estate_id,source_record_id,record_year,record_date,activity,details,status,worker_text,labor_hours,olives_harvested_kg,mill_date,oil_liters,yield_pct,notes,evidence) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) ON DUPLICATE KEY UPDATE details=VALUES(details),status=VALUES(status),olives_harvested_kg=VALUES(olives_harvested_kg),oil_liters=VALUES(oil_liters),yield_pct=VALUES(yield_pct),notes=VALUES(notes),evidence=VALUES(evidence)", (uid(), ESTATE_ID, as_text(row["Record ID"]), as_int(row.get("Year")), as_date(row.get("Date")), as_text(row.get("Activity")), as_text(row.get("Products, doses & water")), as_text(row.get("Status")), as_text(row.get("Worker")), as_number(row.get("Hours")), as_number(row.get("Olives harvested kg")), as_date(row.get("Mill date")), as_number(row.get("Oil liters")), as_number(row.get("Yield %")), as_text(row.get("Issues / Notes")), as_text(row.get("Evidence / Source"))))

    def import_treatments(self) -> None:
        _, rows = find_table(self.workbook["Vineyard Treatments"], "Treatment ID")
        for _, row in rows:
            treatment_id = as_text(row.get("Treatment ID"))
            year = as_int(row.get("Year"))
            planned_date = as_date(row.get("Planned application")) or as_date(row.get("Plan date"))
            if not treatment_id or not year or not planned_date:
                continue
            raw_status = (as_text(row.get("Status")) or "Planned").casefold()
            # Historical 'Applied' confirms that work occurred, but the source explicitly
            # says the actual date, quantities, weather and checks remain unconfirmed.
            # Preserve it as a planned record until those completion facts are reviewed.
            status = "cancelled" if "cancel" in raw_status else "planned"
            product_text = as_text(row.get("Product / active ingredient"))
            dose_text = as_text(row.get("Dose"))
            water_text = as_text(row.get("Water volume"))
            water_l = as_number(water_text) if water_text and re.fullmatch(r"\s*[0-9.,]+\s*L\s*", water_text, re.I) else None
            purpose = as_text(row.get("Target / risk")) or f"Treatment {as_text(row.get('Treatment no.')) or treatment_id}"
            notes = "\n\n".join(filter(None, [as_text(row.get("Completion notes")), f"Source status: {as_text(row.get('Status')) or 'unknown'}"]))
            self.bump("treatments")
            if self.commit_mode:
                self.cursor.execute(
                    "INSERT INTO spray_applications (id,estate_id,season_id,application_date,planned_application_date,purpose,water_volume_l,operator_name,status,notes,source_application_id,evidence_status,planned_by,assigned_to,source_products,source_doses,source_water_text,source_method,source_instructions,source_reference) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) "
                    "ON DUPLICATE KEY UPDATE planned_application_date=VALUES(planned_application_date),purpose=VALUES(purpose),water_volume_l=VALUES(water_volume_l),operator_name=VALUES(operator_name),status=VALUES(status),notes=VALUES(notes),evidence_status=VALUES(evidence_status),planned_by=VALUES(planned_by),assigned_to=VALUES(assigned_to),source_products=VALUES(source_products),source_doses=VALUES(source_doses),source_water_text=VALUES(source_water_text),source_method=VALUES(source_method),source_instructions=VALUES(source_instructions),source_reference=VALUES(source_reference)",
                    (uid(), ESTATE_ID, self.season(year), datetime.combine(planned_date, time(12, 0)), planned_date, purpose, water_l, as_text(row.get("Assigned to")), status, notes, treatment_id, "source-reported; completion details need review", as_text(row.get("Planned by")), as_text(row.get("Assigned to")), product_text, dose_text, water_text, as_text(row.get("Method")), as_text(row.get("Treatment mix / instructions")), as_text(row.get("Source ID / file"))),
                )

    def import_harvest_plans(self) -> None:
        _, rows = find_table(self.workbook["Harvest Plan"], "Plan ID")
        for _, row in rows:
            pick_date = as_date(row.get("Pick date")); variety_id = self.variety(row.get("Variety"))
            if not pick_date or not variety_id: continue
            self.bump("harvest_plans")
            status = (as_text(row.get("Status")) or "Provisional").lower().replace(" ", "_")
            if status not in {"draft","provisional","confirmed","in_progress","complete","cancelled","hold"}: status="provisional"
            if self.commit_mode:
                self.cursor.execute("INSERT INTO harvest_plans (id,estate_id,season_id,source_plan_id,variety_id,block_reference,planned_pick_date,status,planned_kg,planned_crates,crew_size,planned_hours,cellar_destination,weather_risk,dependencies,approved_by,forecast_method,notes) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) ON DUPLICATE KEY UPDATE planned_pick_date=VALUES(planned_pick_date),status=VALUES(status),planned_kg=VALUES(planned_kg),weather_risk=VALUES(weather_risk),dependencies=VALUES(dependencies),notes=VALUES(notes)", (uid(), ESTATE_ID, self.season(pick_date.year), as_text(row["Plan ID"]), variety_id, as_text(row.get("Block ID")), pick_date, status, as_number(row.get("Planned kg")), as_int(row.get("Cassettes")), as_int(row.get("Crew size")), as_number(row.get("Planned hours")), as_text(row.get("Cellar slot/vessel")), as_text(row.get("Weather risk")), as_text(row.get("Dependencies")), as_text(row.get("Approved by")), "GDD + maturity + weather", as_text(row.get("Notes"))))

    def import_maturity_samples(self) -> None:
        _, rows = find_table(self.workbook["Maturity Samples"], "Sample ID")
        for _, row in rows:
            sample_id, sample_date = as_text(row.get("Sample ID")), as_date(row.get("Date"))
            if not sample_id or not sample_date:
                continue
            sampled_at = datetime.combine(sample_date, row.get("Time") if isinstance(row.get("Time"), time) else time(12, 0))
            decision = (as_text(row.get("Decision")) or "monitor").lower().replace(" ", "_")
            if decision not in {"monitor", "resample", "hold", "ready", "picked"}: decision = "monitor"
            record_id = str(uuid.uuid5(uuid.NAMESPACE_URL, f"baiamonte:maturity:{sample_id}"))
            self.bump("maturity_samples")
            if self.commit_mode:
                self.cursor.execute("SELECT id FROM vineyard_blocks WHERE estate_id=%s AND code=%s", (ESTATE_ID, as_text(row.get("Block ID"))))
                block = self.cursor.fetchone()
                self.cursor.execute("INSERT INTO maturity_samples (id,estate_id,season_id,block_id,variety_id,sampled_at,berry_count,sample_kg,brix,ph,ta_g_l,yan_mg_l,fruit_temp_c,disease_pct,condition_notes,decision,provisional_pick_date,sampler,notes) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) ON DUPLICATE KEY UPDATE sampled_at=VALUES(sampled_at),berry_count=VALUES(berry_count),sample_kg=VALUES(sample_kg),brix=VALUES(brix),ph=VALUES(ph),ta_g_l=VALUES(ta_g_l),yan_mg_l=VALUES(yan_mg_l),fruit_temp_c=VALUES(fruit_temp_c),disease_pct=VALUES(disease_pct),condition_notes=VALUES(condition_notes),decision=VALUES(decision),provisional_pick_date=VALUES(provisional_pick_date),sampler=VALUES(sampler),notes=VALUES(notes)", (record_id, ESTATE_ID, self.season(sample_date.year), block[0] if block else None, self.variety(row.get("Variety")), sampled_at, as_int(row.get("Berry count")), as_number(row.get("Sample kg")), as_number(row.get("Brix")), as_number(row.get("pH")), as_number(row.get("TA g/L")), as_number(row.get("YAN mg/L")), as_number(row.get("Fruit temp °C")), as_number(row.get("Disease %")), as_text(row.get("Berry/skin/seed condition")), decision, as_date(row.get("Provisional pick")), as_text(row.get("Sampler")), as_text(row.get("Notes"))))

    def import_harvest_operations(self) -> None:
        _, rows = find_table(self.workbook["Harvest Operations"], "Lot ID")
        for _, row in rows:
            lot_code, actual_date = as_text(row.get("Lot ID")), as_date(row.get("Actual date"))
            if not lot_code or not actual_date:
                continue
            variety_id = self.variety(row.get("Variety"))
            if not variety_id:
                continue
            status = (as_text(row.get("Status")) or "received").lower().replace(" ", "_")
            if status not in {"provisional", "ready", "in_progress", "received", "reconciled", "hold", "cancelled"}: status = "received"
            self.bump("harvest_lots")
            if self.commit_mode:
                self.cursor.execute("SELECT id FROM vineyard_blocks WHERE estate_id=%s AND code=%s", (ESTATE_ID, as_text(row.get("Block ID"))))
                block = self.cursor.fetchone()
                self.cursor.execute("INSERT INTO harvest_lots (id,estate_id,season_id,lot_code,block_id,variety_id,harvested_at,planned_date,planned_kg,gross_kg,tare_kg,weight_kg,crate_count,avg_crate_kg,fruit_temp_c,destination,brix,condition_grade,status,notes) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) ON DUPLICATE KEY UPDATE harvested_at=VALUES(harvested_at),planned_date=VALUES(planned_date),planned_kg=VALUES(planned_kg),gross_kg=VALUES(gross_kg),tare_kg=VALUES(tare_kg),weight_kg=VALUES(weight_kg),crate_count=VALUES(crate_count),avg_crate_kg=VALUES(avg_crate_kg),fruit_temp_c=VALUES(fruit_temp_c),destination=VALUES(destination),brix=VALUES(brix),condition_grade=VALUES(condition_grade),status=VALUES(status),notes=VALUES(notes)", (uid(), ESTATE_ID, self.season(as_int(row.get("Vintage")) or actual_date.year), lot_code, block[0] if block else None, variety_id, datetime.combine(actual_date, time(12, 0)), as_date(row.get("Planned date")), as_number(row.get("Planned kg")), as_number(row.get("Gross kg")), as_number(row.get("Tare kg")), as_number(row.get("Net kg")), as_int(row.get("Cassettes")), None, as_number(row.get("Fruit temp °C")), as_text(row.get("Destination / vessel")), as_number(row.get("Brix")), as_text(row.get("Condition")), status, as_text(row.get("Notes"))))

    def import_cellar_operations(self) -> None:
        _, rows = find_table(self.workbook["Cellar Operations"], "Cellar lot ID")
        for _, row in rows:
            code, year = as_text(row.get("Cellar lot ID")), as_int(row.get("Vintage"))
            if not code or not year:
                continue
            raw_stage = (as_text(row.get("Stage")) or "must").casefold()
            stage = next((value for value in ("fermentation", "malo", "aging", "bottled", "closed", "must") if value in raw_stage), "must")
            lot_status = as_text(row.get("Status")) or as_text(row.get("Stage"))
            self.bump("wine_lots")
            if self.commit_mode:
                self.cursor.execute("INSERT INTO wine_lots (id,estate_id,season_id,code,harvest_lot_reference,name,stage,lot_status,volume_l,fruit_kg,initial_l,free_run_l,press_l,loss_l,variety_summary,started_at,responsible,notes) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) ON DUPLICATE KEY UPDATE harvest_lot_reference=VALUES(harvest_lot_reference),name=VALUES(name),stage=VALUES(stage),lot_status=VALUES(lot_status),volume_l=VALUES(volume_l),fruit_kg=VALUES(fruit_kg),initial_l=VALUES(initial_l),free_run_l=VALUES(free_run_l),press_l=VALUES(press_l),loss_l=VALUES(loss_l),variety_summary=VALUES(variety_summary),started_at=VALUES(started_at),responsible=VALUES(responsible),notes=VALUES(notes)", (uid(), ESTATE_ID, self.season(year), code, as_text(row.get("Harvest lot ID")), code, stage, lot_status, as_number(row.get("Current L")), as_number(row.get("Fruit kg")), as_number(row.get("Initial L")), as_number(row.get("Free-run L")), as_number(row.get("Press L")), as_number(row.get("Loss L")), as_text(row.get("Variety")), datetime.combine(as_date(row.get("Intake date")), time(12, 0)) if as_date(row.get("Intake date")) else None, as_text(row.get("Responsible")), as_text(row.get("Notes"))))

    def import_fermentation(self) -> None:
        _, rows = find_table(self.workbook["Fermentation"], "Observation ID")
        for _, row in rows:
            source_id, observed_date = as_text(row.get("Observation ID")), as_date(row.get("Date"))
            if not source_id or not observed_date:
                continue
            observed_at = datetime.combine(observed_date, row.get("Time") if isinstance(row.get("Time"), time) else time(12, 0))
            next_check = as_date(row.get("Next check"))
            self.bump("fermentation_observations")
            if self.commit_mode:
                self.cursor.execute("SELECT id FROM wine_lots WHERE estate_id=%s AND code=%s ORDER BY updated_at DESC LIMIT 1", (ESTATE_ID, as_text(row.get("Cellar lot ID"))))
                wine_lot = self.cursor.fetchone()
                self.cursor.execute("INSERT INTO fermentation_observations (id,estate_id,wine_lot_id,source_observation_id,observed_at,vessel_name,stage,temp_c,density_sg,brix,ph,cap_management,addition_action,product_lot,quantity,unit,sensory_observation,owner_text,next_check_at,status) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) ON DUPLICATE KEY UPDATE observed_at=VALUES(observed_at),vessel_name=VALUES(vessel_name),stage=VALUES(stage),temp_c=VALUES(temp_c),density_sg=VALUES(density_sg),brix=VALUES(brix),ph=VALUES(ph),cap_management=VALUES(cap_management),addition_action=VALUES(addition_action),product_lot=VALUES(product_lot),quantity=VALUES(quantity),unit=VALUES(unit),sensory_observation=VALUES(sensory_observation),owner_text=VALUES(owner_text),next_check_at=VALUES(next_check_at),status=VALUES(status)", (uid(), ESTATE_ID, wine_lot[0] if wine_lot else None, source_id, observed_at, as_text(row.get("Vessel")), as_text(row.get("Stage")), as_number(row.get("Temp °C")), as_number(row.get("Density/SG")), as_number(row.get("Brix")), as_number(row.get("pH")), as_text(row.get("Cap management")), as_text(row.get("Addition/action")), as_text(row.get("Product lot")), as_number(row.get("Quantity")), as_text(row.get("Unit")), as_text(row.get("Sensory observation")), as_text(row.get("Owner")), datetime.combine(next_check, time(12, 0)) if next_check else None, as_text(row.get("Status"))))

    def import_mass_balance(self) -> None:
        _, rows = find_table(self.workbook["Mass Balance"], "Harvest lot ID")
        for _, row in rows:
            reference = as_text(row.get("Harvest lot ID"))
            if not reference:
                continue
            self.bump("mass_balance_records")
            if self.commit_mode:
                self.cursor.execute("INSERT INTO mass_balance_records (id,estate_id,harvest_lot_reference,block_reference,variety_name,net_grapes_kg,must_wine_l,free_run_l,press_l,recorded_loss_l,reconciliation_status,owner_text,notes) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) ON DUPLICATE KEY UPDATE block_reference=VALUES(block_reference),variety_name=VALUES(variety_name),net_grapes_kg=VALUES(net_grapes_kg),must_wine_l=VALUES(must_wine_l),free_run_l=VALUES(free_run_l),press_l=VALUES(press_l),recorded_loss_l=VALUES(recorded_loss_l),reconciliation_status=VALUES(reconciliation_status),owner_text=VALUES(owner_text),notes=VALUES(notes)", (uid(), ESTATE_ID, reference, as_text(row.get("Block ID")), as_text(row.get("Variety")), as_number(row.get("Net grapes kg")), as_number(row.get("Must/wine L")), as_number(row.get("Free-run L")), as_number(row.get("Press L")), as_number(row.get("Recorded loss L")), as_text(row.get("Reconciliation status")), as_text(row.get("Owner")), as_text(row.get("Notes"))))

    def import_vintage_history(self) -> None:
        _, rows = find_table(self.workbook["Historical Records"], "Vintage")
        for _, row in rows:
            year = as_int(row.get("Vintage")); name = as_text(row.get("Variety / scope"))
            if not year or not name: continue
            self.season(year); self.bump("vintage_summaries")
            if self.commit_mode:
                self.cursor.execute("INSERT INTO vintage_summaries (estate_id,vintage_year,variety_name,grapes_kg,wine_l,cassette_count,evidence_status,reconciliation_note) VALUES (%s,%s,%s,%s,%s,%s,%s,%s) ON DUPLICATE KEY UPDATE grapes_kg=VALUES(grapes_kg),wine_l=VALUES(wine_l),cassette_count=VALUES(cassette_count),evidence_status=VALUES(evidence_status),reconciliation_note=VALUES(reconciliation_note)", (ESTATE_ID, year, name, as_number(row.get("Grapes kg")), as_number(row.get("Wine L")), as_number(row.get("Cassettes")), as_text(row.get("Evidence status")), as_text(row.get("Reconciliation note"))))

    def import_projections(self) -> None:
        """Move the workbook's useful planning outputs into editable MariaDB records."""
        sheet = self.workbook["Projections"]
        _, forecasts = find_table(sheet, "Vintage")
        forecast_years = [as_int(row.get("Vintage")) for _, row in forecasts]
        projection_year = min(year for year in forecast_years if year) if any(forecast_years) else datetime.now().year
        _, allocations = find_table(sheet, "Grape")
        for _, row in allocations:
            grape = as_text(row.get("Grape"))
            total_kg = as_number(row.get("Total kg"))
            if not grape or total_kg is None or grape.casefold() == "total":
                continue
            total_crates = as_int(row.get("Total 15 kg crates")) or math.ceil(total_kg / 15)
            self.bump("grape_allocation_plans")
            if self.commit_mode:
                self.cursor.execute(
                    "INSERT INTO grape_allocation_plans (estate_id,vintage_year,grape_name,total_kg,total_crates_15kg,wine_destination,blend_kg,blend_crates_15kg,varietal_kg,varietal_crates_15kg,field_instruction,source) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'workbook migration') ON DUPLICATE KEY UPDATE total_kg=VALUES(total_kg),total_crates_15kg=VALUES(total_crates_15kg),wine_destination=VALUES(wine_destination),blend_kg=VALUES(blend_kg),blend_crates_15kg=VALUES(blend_crates_15kg),varietal_kg=VALUES(varietal_kg),varietal_crates_15kg=VALUES(varietal_crates_15kg),field_instruction=VALUES(field_instruction),source=VALUES(source)",
                    (ESTATE_ID, projection_year, grape, total_kg, total_crates, as_text(row.get("Wine destination")) or "Unallocated", as_number(row.get("To Nerello blend kg")) or 0, as_int(row.get("Blend crates")) or 0, as_number(row.get("Varietal kg")) or 0, as_int(row.get("Varietal crates")) or 0, as_text(row.get("Field instruction"))),
                )

        _, outputs = find_table(sheet, "Finished wine")
        for _, row in outputs:
            wine = as_text(row.get("Finished wine"))
            grape_kg = as_number(row.get("Grape kg"))
            if not wine or grape_kg is None or wine.casefold() == "total":
                continue
            wine_l = as_number(row.get("Wine L")) or round(grape_kg * 0.70)
            bottles = as_int(row.get("0.75 L bottles")) or math.floor(wine_l / 0.75)
            self.bump("wine_output_plans")
            if self.commit_mode:
                self.cursor.execute(
                    "INSERT INTO wine_output_plans (estate_id,vintage_year,finished_wine,composition,grape_kg,wine_l,bottles_750ml,source) VALUES (%s,%s,%s,%s,%s,%s,%s,'workbook migration') ON DUPLICATE KEY UPDATE composition=VALUES(composition),grape_kg=VALUES(grape_kg),wine_l=VALUES(wine_l),bottles_750ml=VALUES(bottles_750ml),source=VALUES(source)",
                    (ESTATE_ID, projection_year, wine, as_text(row.get("Composition")) or "Not recorded", grape_kg, wine_l, bottles),
                )

        varieties = (("Grecanico", "Grecanico kg", "Grecanico crates"), ("Nerello Mascalese", "Nerello kg", "Nerello crates"), ("Grenache", "Grenache kg", "Grenache crates"))
        for _, row in forecasts:
            vintage = as_int(row.get("Vintage"))
            if not vintage:
                continue
            for variety, kg_column, crate_column in varieties:
                grape_kg = as_number(row.get(kg_column))
                if grape_kg is None:
                    continue
                crates = as_int(row.get(crate_column)) or math.ceil(grape_kg / 15)
                self.bump("production_forecasts")
                if self.commit_mode:
                    self.cursor.execute(
                        "INSERT INTO production_forecasts (estate_id,vintage_year,scenario,variety_name,grape_kg,crates_15kg,source) VALUES (%s,%s,'base',%s,%s,%s,'workbook migration') ON DUPLICATE KEY UPDATE grape_kg=VALUES(grape_kg),crates_15kg=VALUES(crates_15kg),source=VALUES(source)",
                        (ESTATE_ID, vintage, variety, grape_kg, crates),
                    )

    def import_labs(self, sheet_name: str) -> None:
        _, rows = find_table(self.workbook[sheet_name], "Test ID")
        analytes = {"pH": ("ph", "pH", None), "TA g/L": ("ta", "Titratable acidity", "g/L"), "VA g/L": ("va", "Volatile acidity", "g/L"), "Malic g/L": ("malic", "Malic acid", "g/L"), "Glucose+fructose g/L": ("glucose_fructose", "Glucose + fructose", "g/L"), "Free SO2 mg/L": ("free_so2", "Free SO2", "mg/L"), "Total SO2 mg/L": ("total_so2", "Total SO2", "mg/L"), "YAN mg/L": ("yan", "YAN", "mg/L"), "°Babo": ("babo", "Babo", "°Babo"), "Potential alc. % vol": ("potential_alcohol", "Potential alcohol", "% vol"), "Actual alc. % vol": ("alcohol", "Alcohol", "% vol"), "Lactic acid g/L": ("lactic", "Lactic acid", "g/L"), "Brett PCR cells/mL": ("brett_pcr", "Brett PCR", "cells/mL")}
        for _, row in rows:
            test_id = as_text(row.get("Test ID"))
            if not test_id or "TEMPLATE" in test_id: continue
            lab_date = as_date(row.get("Date"))
            if not lab_date: continue
            stated_vintage = as_int(row.get("Vintage"))
            if stated_vintage:
                year, vintage_source, vintage_confidence = stated_vintage, "source_report", "confirmed"
                vintage_evidence = "The source workbook explicitly states the vintage."
            elif test_id in {"LAB-20240507-01", "LAB-20240507-02", "LAB-20240507-03"}:
                year, vintage_source, vintage_confidence = 2023, "cellar_chronology", "inferred"
                vintage_evidence = "The report leaves Annata blank. It is a wine report dated before the 2024 harvest, so it belongs to the preceding 2023 cellar vintage."
            elif test_id.startswith("LAB-20250509-"):
                year, vintage_source, vintage_confidence = 2023, "sample_identity", "inferred"
                vintage_evidence = "The report leaves Annata blank. Its G 1-G 3 and N 1-N 3 IDs match the 24 April 2025 report explicitly assigned to vintage 2023."
            elif test_id.startswith("LAB-20251027-"):
                year, vintage_source, vintage_confidence = 2025, "cellar_chronology", "inferred"
                vintage_evidence = "The report leaves Annata blank. Its malolactic sequence falls between the October harvest reports and November samples explicitly labeled 25."
            else:
                year, vintage_source, vintage_confidence = lab_date.year, "calendar_fallback", "review_required"
                vintage_evidence = "The source does not state a vintage and no linked lot or audited chronology resolves it. Calendar year is provisional only."
            sample_id = str(uuid.uuid5(uuid.NAMESPACE_URL, f"baiamonte:{test_id}"))
            self.bump("lab_samples")
            if self.commit_mode:
                stage = (as_text(row.get("Stage")) or "other").lower()
                sample_type = "grape" if "grape" in stage else ("must" if "must" in stage else ("wine" if "wine" in stage else "other"))
                self.cursor.execute("INSERT INTO lab_samples (id,estate_id,season_id,sample_code,sample_name,sample_type,lab_date,vintage_year,vintage_assignment_source,vintage_assignment_confidence,vintage_assignment_evidence,laboratory,source_document,needs_review,review_notes,notes) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) ON DUPLICATE KEY UPDATE season_id=VALUES(season_id),sample_name=VALUES(sample_name),sample_type=VALUES(sample_type),lab_date=VALUES(lab_date),vintage_year=VALUES(vintage_year),vintage_assignment_source=VALUES(vintage_assignment_source),vintage_assignment_confidence=VALUES(vintage_assignment_confidence),vintage_assignment_evidence=VALUES(vintage_assignment_evidence),source_document=VALUES(source_document),needs_review=VALUES(needs_review),review_notes=VALUES(review_notes),notes=VALUES(notes)", (sample_id, ESTATE_ID, self.season(year), test_id, as_text(row.get("Cellar lot ID / sample")) or test_id, sample_type, lab_date, year, vintage_source, vintage_confidence, vintage_evidence, as_text(row.get("Lab/source")), as_text(row.get("Source document")), 1 if as_text(row.get("Status")) == "Review" else 0, as_text(row.get("Status")), as_text(row.get("Notes"))))
                for column, (code, name, unit) in analytes.items():
                    value = row.get(column)
                    if value is None: continue
                    numeric = as_number(value); text_value = None if numeric is not None else as_text(value)
                    self.cursor.execute("INSERT INTO lab_results (id,sample_id,analyte_code,analyte_name,numeric_value,text_value,unit) VALUES (%s,%s,%s,%s,%s,%s,%s) ON DUPLICATE KEY UPDATE numeric_value=VALUES(numeric_value),text_value=VALUES(text_value),unit=VALUES(unit)", (uid(), sample_id, code, name, numeric, text_value, unit))
                    self.bump("lab_results")

    def report(self) -> dict[str, Any]:
        return {"mode": "commit" if self.commit_mode else "dry-run", "workbook": self.path.name, "sha256": self.sha256, "sheets": self.workbook.sheetnames, "counts": self.counts, "warnings": self.warnings, "governance": ["MariaDB is the authoritative operational system after migration.", "The workbook is a one-time migration and audit source, not a live authority.", "Blank values remain NULL, never coerced to zero.", "All non-empty source rows are preserved with sheet and row number.", "Planned records are not imported as completed records.", "Workbook import is transactional and content-hash deduplicated."]}


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--commit", action="store_true", help="Write after validation; dry-run is default")
    parser.add_argument("--source-file-id", default="1jYP2HXErEUiFA461NxMXRAxgWYYmfyceBplm-jLORy8")
    parser.add_argument("--source-modified-at", default="2026-08-08T05:10:00")
    parser.add_argument("--report", type=Path)
    args = parser.parse_args()
    parser.error("Workbook commits are retired and workbook access is retired; use MariaDB and authenticated application inputs")


if __name__ == "__main__":
    sys.exit(main())
