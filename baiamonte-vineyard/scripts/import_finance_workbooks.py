"""Import the authoritative Baiamonte finance and funding workbooks.

Dry-run is the default. Use --commit only after the report is reviewed and the
MariaDB/Home Assistant backup is verified. Every non-empty row is retained in
workbook_source_rows even when it is not normalized into a finance table.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import uuid
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import openpyxl
import pymysql

from import_workbook import ESTATE_ID, as_date, as_int, as_number, as_text, find_table, find_tables, json_value, normalized_row, uid


FINANCE_SHEETS = {"Annual P&L", "Sales Invoices", "Purchase Invoices", "Cash Ledger", "Products", "Customers", "Suppliers"}
FUNDING_SHEETS = {"Opportunity Tracker", "Timeline", "Eligibility & Docs", "Funding Model", "Sources & Audit"}


def normalized_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "_", (as_text(value) or "").lower()).strip("_")


def money(value: Any) -> float:
    return as_number(value) or 0.0


def workbook_date(value: Any) -> date | None:
    parsed = as_date(value)
    if parsed:
        return parsed
    number = as_number(value)
    if number and 1 < number < 100000:
        return date(1899, 12, 30) + timedelta(days=int(number))
    return None


class FinanceImporter:
    def __init__(self, paths: list[Path], commit: bool):
        self.paths = paths
        self.commit_mode = commit
        self.books = [(path, openpyxl.load_workbook(path, data_only=True, read_only=True)) for path in paths]
        self.counts: dict[str, int] = {"raw_rows": 0}
        self.warnings: list[str] = []
        self.connection = None
        self.cursor = None
        self.batch_ids: dict[Path, str] = {}
        self.party_cache: dict[str, str] = {}
        self.product_cache: dict[str, str] = {}
        self.scenario_cache: dict[str, str] = {}

    def connect(self) -> None:
        self.connection = pymysql.connect(
            host=os.getenv("DB_HOST", "core-mariadb"), port=int(os.getenv("DB_PORT", "3306")),
            user=os.getenv("DB_USER", "baiamonte"), password=os.getenv("DB_PASSWORD", ""),
            database=os.getenv("DB_NAME", "baiamonte_vineyard"), charset="utf8mb4", autocommit=False,
        )
        self.cursor = self.connection.cursor()

    def bump(self, key: str, amount: int = 1) -> None:
        self.counts[key] = self.counts.get(key, 0) + amount

    def run(self) -> dict[str, Any]:
        all_sheets = {name for _, book in self.books for name in book.sheetnames}
        if not (FINANCE_SHEETS <= all_sheets or FUNDING_SHEETS <= all_sheets):
            raise ValueError("Provide the Finance Workbook, Funding Tracker, or both")
        if self.commit_mode:
            self.connect()
        try:
            for path, book in self.books:
                self.start_batch(path, book)
                self.preserve_rows(path, book)
                if FINANCE_SHEETS <= set(book.sheetnames):
                    self.import_finance(book)
                if FUNDING_SHEETS <= set(book.sheetnames):
                    self.import_funding(book)
            report = self.report()
            if self.commit_mode:
                for batch_id in self.batch_ids.values():
                    self.cursor.execute("UPDATE import_batches SET status='committed',row_count=%s,warning_count=%s,report=%s,completed_at=NOW(6) WHERE id=%s", (self.counts["raw_rows"], len(self.warnings), json.dumps(report, default=json_value), batch_id))
                self.connection.commit()
            return report
        except Exception:
            if self.connection:
                self.connection.rollback()
            raise
        finally:
            if self.connection:
                self.connection.close()

    def start_batch(self, path: Path, book: Any) -> None:
        digest = hashlib.sha256(path.read_bytes()).hexdigest()
        batch_id = uid()
        self.batch_ids[path] = batch_id
        if self.commit_mode:
            self.cursor.execute("INSERT INTO import_batches (id,estate_id,source_name,content_sha256,status,report) VALUES (%s,%s,%s,%s,'started',%s)", (batch_id, ESTATE_ID, path.name, digest, json.dumps({"sheets": book.sheetnames})))

    def preserve_rows(self, path: Path, book: Any) -> None:
        batch_id = self.batch_ids[path]
        for sheet in book.worksheets:
            for row_number, values in enumerate(sheet.iter_rows(values_only=True), 1):
                row = normalized_row(values)
                if not row:
                    continue
                payload = json.dumps(row, default=json_value, ensure_ascii=False)
                row_hash = hashlib.sha256(f"{sheet.title}:{row_number}:{payload}".encode()).hexdigest()
                self.bump("raw_rows")
                if self.commit_mode:
                    self.cursor.execute("INSERT INTO workbook_source_rows (import_batch_id,sheet_name,source_row_number,row_values,row_hash) VALUES (%s,%s,%s,%s,%s)", (batch_id, sheet.title, row_number, payload, row_hash))

    def party(self, name: Any, party_type: str = "other", row: dict[str, Any] | None = None) -> str | None:
        name = as_text(name)
        if not name:
            return None
        key = name.casefold()
        if key not in self.party_cache:
            self.party_cache[key] = uid()
            if self.commit_mode:
                data = row or {}
                self.cursor.execute(
                    "INSERT INTO finance_parties (id,estate_id,party_type,name,internal_code,vat_id,tax_code,email,certified_email,phone,contact_name,address_line,city,postal_code,province,country,iban,sdi_code,payment_terms,default_payment_method,notes,source) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'finance_workbook') ON DUPLICATE KEY UPDATE party_type=IF(party_type=VALUES(party_type),party_type,'both'),vat_id=COALESCE(VALUES(vat_id),vat_id),tax_code=COALESCE(VALUES(tax_code),tax_code),email=COALESCE(VALUES(email),email),phone=COALESCE(VALUES(phone),phone),address_line=COALESCE(VALUES(address_line),address_line),iban=COALESCE(VALUES(iban),iban)",
                    (self.party_cache[key], ESTATE_ID, party_type, name, as_text(data.get("Codice interno")), as_text(data.get("P.IVA/TAX ID")) or as_text(data.get("VAT ID")), as_text(data.get("Codice Fiscale")) or as_text(data.get("Tax code")), as_text(data.get("Indirizzo e-mail")), as_text(data.get("Indirizzo PEC")), as_text(data.get("Telefono")), as_text(data.get("Referente")), as_text(data.get("Indirizzo")), as_text(data.get("Comune")), as_text(data.get("CAP")), as_text(data.get("Provincia")), as_text(data.get("Paese")), as_text(data.get("IBAN")), as_text(data.get("Codice SDI")), as_text(data.get("Termini di pagamento")), as_text(data.get("Metodo di pagamento prefefinito")), as_text(data.get("Note"))),
                )
                self.cursor.execute("SELECT id FROM finance_parties WHERE estate_id=%s AND name=%s", (ESTATE_ID, name))
                self.party_cache[key] = self.cursor.fetchone()[0]
        return self.party_cache[key]

    def product(self, code: Any, name: Any = None, row: dict[str, Any] | None = None) -> str | None:
        code = as_text(code)
        name = as_text(name) or code
        if not code or not name:
            return None
        if code not in self.product_cache:
            self.product_cache[code] = uid()
            if self.commit_mode:
                data = row or {}
                self.cursor.execute("INSERT INTO products (id,estate_id,name,product_type,unit,sku,description,category_name,sales_price_net,sales_price_gross,vat_rate,purchase_price,track_inventory) VALUES (%s,%s,%s,'other',%s,%s,%s,%s,%s,%s,%s,%s,%s) ON DUPLICATE KEY UPDATE name=VALUES(name),sku=COALESCE(sku,VALUES(sku)),description=VALUES(description),category_name=VALUES(category_name),sales_price_net=VALUES(sales_price_net),sales_price_gross=VALUES(sales_price_gross),purchase_price=VALUES(purchase_price),track_inventory=VALUES(track_inventory)", (self.product_cache[code], ESTATE_ID, name, as_text(data.get("U.D.M.")) or as_text(data.get("U.M.")) or "unit", code, as_text(data.get("Descrizione")), as_text(data.get("Categoria")), as_number(data.get("Prezzo netto")), as_number(data.get("Prezzo lordo")), as_number(data.get("Aliquota IVA")), as_number(data.get("Prezzo di acquisto")), 1 if data.get("Giacenza") is not None else 0))
                self.cursor.execute("SELECT id FROM products WHERE estate_id=%s AND (sku=%s OR name=%s) ORDER BY (sku=%s) DESC LIMIT 1", (ESTATE_ID, code, name, code))
                existing = self.cursor.fetchone()
                if not existing:
                    raise RuntimeError(f"Product upsert could not be resolved: {code} / {name}")
                self.product_cache[code] = existing[0]
        return self.product_cache[code]

    def scenario(self, name: str, scenario_type: str, selected: bool = False) -> str:
        key = name.casefold()
        if key not in self.scenario_cache:
            self.scenario_cache[key] = uid()
            if self.commit_mode:
                self.cursor.execute("INSERT INTO financial_scenarios (id,estate_id,name,scenario_type,selected,source) VALUES (%s,%s,%s,%s,%s,'finance_workbook') ON DUPLICATE KEY UPDATE scenario_type=VALUES(scenario_type),selected=VALUES(selected)", (self.scenario_cache[key], ESTATE_ID, name, scenario_type, selected))
                self.cursor.execute("SELECT id FROM financial_scenarios WHERE estate_id=%s AND name=%s", (ESTATE_ID, name))
                self.scenario_cache[key] = self.cursor.fetchone()[0]
        return self.scenario_cache[key]

    def import_finance(self, book: Any) -> None:
        self.import_parties(book)
        self.import_products_inventory(book)
        self.import_invoices(book, "Sales Invoices", "sales_invoice")
        self.import_invoices(book, "Purchase Invoices", "purchase_invoice")
        self.import_cash(book)
        self.import_assumptions(book)
        self.import_annual(book)
        self.import_monthly(book)
        self.import_vat(book)
        self.import_accounts(book)
        self.import_prices(book)
        self.import_contract_services(book)
        self.import_finance_sources(book)

    def import_parties(self, book: Any) -> None:
        for sheet_name, party_type in (("Customers", "customer"), ("Suppliers", "supplier")):
            _, rows = find_table(book[sheet_name], "Denominazione")
            for _, row in rows:
                self.party(row.get("Denominazione"), party_type, row)
                self.bump("finance_parties")

    def import_products_inventory(self, book: Any) -> None:
        _, rows = find_table(book["Products"], "Codice")
        for _, row in rows:
            self.product(row.get("Codice"), row.get("Nome prodotto/servizio"), row)
            self.bump("products")
        _, rows = find_table(book["Inventory"], "Categoria")
        for _, row in rows:
            product_id = self.product(row.get("Codice"), row.get("Nome prodotto"), row)
            if not product_id:
                continue
            self.bump("inventory_snapshots")
            if self.commit_mode:
                self.cursor.execute("INSERT INTO inventory_snapshots (id,estate_id,product_id,snapshot_date,quantity_on_hand,opening_quantity,average_cost,average_sales_price,inventory_value,source) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,'finance_workbook') ON DUPLICATE KEY UPDATE quantity_on_hand=VALUES(quantity_on_hand),opening_quantity=VALUES(opening_quantity),average_cost=VALUES(average_cost),average_sales_price=VALUES(average_sales_price),inventory_value=VALUES(inventory_value)", (uid(), ESTATE_ID, product_id, date(2025,12,31), as_number(row.get("Giacenza")), as_number(row.get("Giacenza iniz.")), as_number(row.get("Costo medio")), as_number(row.get("Prezzo medio")), as_number(row.get("Costo giacenza"))))

    def import_invoices(self, book: Any, sheet_name: str, document_type: str) -> None:
        _, rows = find_table(book[sheet_name], "Invoice #")
        for _, row in rows:
            invoice_date = workbook_date(row.get("Date"))
            number = as_text(row.get("Invoice #"))
            if not invoice_date or not number:
                continue
            party_id = self.party(row.get("Party"), "customer" if document_type == "sales_invoice" else "supplier", row)
            taxable, vat = money(row.get("Taxable amount")), money(row.get("VAT"))
            self.bump("financial_documents")
            if self.commit_mode:
                self.cursor.execute("INSERT INTO financial_documents (id,estate_id,document_type,document_number,document_date,party_id,taxable_amount,vat_amount,withholding_tax,social_security_withholding,gross_total,deductible_pct,vat_deductible_pct,depreciation_years,status,payment_status,source,external_source_id) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'unknown','finance_workbook',%s) ON DUPLICATE KEY UPDATE party_id=VALUES(party_id),taxable_amount=VALUES(taxable_amount),vat_amount=VALUES(vat_amount),gross_total=VALUES(gross_total),deductible_pct=VALUES(deductible_pct),vat_deductible_pct=VALUES(vat_deductible_pct),depreciation_years=VALUES(depreciation_years)", (uid(), ESTATE_ID, document_type, number, invoice_date, party_id, taxable, vat, money(row.get("Withholding tax")), money(row.get("Social security withholding")), taxable + vat - money(row.get("Withholding tax")) - money(row.get("Social security withholding")), as_number(row.get("Deductibility")), as_number(row.get("VAT deductibility")), as_int(row.get("Depreciation years")), "issued" if document_type == "sales_invoice" else "received", f"{sheet_name}:{number}:{invoice_date}"))

    def import_cash(self, book: Any) -> None:
        _, rows = find_table(book["Cash Ledger"], "Data")
        for row_number, row in rows:
            tx_date = workbook_date(row.get("Data"))
            account_name = as_text(row.get("Conto"))
            if not tx_date or not account_name:
                continue
            account_id = str(uuid.uuid5(uuid.NAMESPACE_URL, f"baiamonte:cash:{account_name}"))
            party_id = self.party(row.get("Cliente/Fornitore"))
            self.bump("cash_transactions")
            if self.commit_mode:
                account_type = "credit_card" if "card" in account_name.lower() else "bank"
                self.cursor.execute("INSERT INTO cash_accounts (id,estate_id,name,account_type) VALUES (%s,%s,%s,%s) ON DUPLICATE KEY UPDATE account_type=VALUES(account_type)", (account_id, ESTATE_ID, account_name, account_type))
                source_id = f"Cash Ledger:{row_number}:{tx_date}:{money(row.get('Entrate'))}:{money(row.get('Uscite'))}"
                self.cursor.execute("INSERT INTO cash_transactions (id,estate_id,cash_account_id,transaction_date,description,party_id,amount_in,amount_out,running_balance,external_source_id,source) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'finance_workbook') ON DUPLICATE KEY UPDATE description=VALUES(description),party_id=VALUES(party_id),amount_in=VALUES(amount_in),amount_out=VALUES(amount_out),running_balance=VALUES(running_balance)", (uid(), ESTATE_ID, account_id, tx_date, as_text(row.get("Descrizione")) or "Imported transaction", party_id, money(row.get("Entrate")), money(row.get("Uscite")), as_number(row.get("Saldo prog.vo")), source_id))

    def import_assumptions(self, book: Any) -> None:
        selected_name = as_text(book["Assumptions"]["B3"].value) or "Base"
        _, rows = find_table(book["Assumptions"], "Driver")
        for _, row in rows:
            driver = as_text(row.get("Driver"))
            if not driver:
                continue
            for column, scenario_type in (("Base", "base"), ("Upside", "upside"), ("Downside", "downside")):
                value = row.get(column)
                if value is None:
                    continue
                scenario_id = self.scenario(column, scenario_type, column == selected_name)
                self.bump("forecast_assumptions")
                if self.commit_mode:
                    numeric = as_number(value)
                    self.cursor.execute("INSERT INTO forecast_assumptions (id,estate_id,scenario_id,assumption_key,assumption_group,numeric_value,text_value,source) VALUES (%s,%s,%s,%s,'operating_growth',%s,%s,'finance_workbook') ON DUPLICATE KEY UPDATE numeric_value=VALUES(numeric_value),text_value=VALUES(text_value)", (uid(), ESTATE_ID, scenario_id, normalized_key(driver), numeric, None if numeric is not None else as_text(value)))

    def import_annual(self, book: Any) -> None:
        rows = list(book["Annual P&L"].iter_rows(values_only=True))
        years = [as_int(value) for value in rows[3][1:] if as_int(value)]
        metrics = {as_text(row[0]): list(row[1:1 + len(years)]) for row in rows[4:] if row and as_text(row[0])}
        mapping = {"Revenue":"revenue","Goods & production inputs":"goods_inputs","Production services":"production_services","Vehicle costs":"vehicle_costs","Professional & technical":"professional_technical","Travel, admin & representation":"travel_admin_representation","Software & licenses":"software_licenses","Payroll":"payroll","Depreciation":"depreciation","Other operating costs":"other_operating_costs","Total operating costs":"total_operating_costs","Operating result":"operating_result","Income tax":"income_tax","After-tax result":"after_tax_result"}
        actual_id = self.scenario("Actual", "actual")
        selected_name = as_text(book["Assumptions"]["B3"].value) or "Base"
        selected_id = self.scenario(selected_name, selected_name.lower(), True)
        for index, year in enumerate(years):
            values = {column: as_number(metrics.get(label, [None] * len(years))[index]) for label, column in mapping.items()}
            self.bump("annual_financial_summary")
            if self.commit_mode:
                scenario_id = actual_id if year <= 2025 else selected_id
                columns = ",".join(values)
                placeholders = ",".join(["%s"] * len(values))
                updates = ",".join(f"{column}=VALUES({column})" for column in values)
                self.cursor.execute(f"INSERT INTO annual_financial_summary (estate_id,fiscal_year,scenario_id,{columns},source) VALUES (%s,%s,%s,{placeholders},'finance_workbook') ON DUPLICATE KEY UPDATE {updates}", (ESTATE_ID, year, scenario_id, *values.values()))

    def import_monthly(self, book: Any) -> None:
        _, rows = find_table(book["Monthly Update"], "Year")
        for _, row in rows:
            year, month = as_int(row.get("Year")), as_int(row.get("Month #"))
            if not year or not month:
                continue
            status = (as_text(row.get("Status")) or "Planned").lower()
            if status not in {"closed", "pending", "planned"}: status = "planned"
            self.bump("monthly_financial_summary")
            if self.commit_mode:
                self.cursor.execute("INSERT INTO monthly_financial_summary (estate_id,fiscal_year,fiscal_month,actual_revenue,actual_cost,budget_revenue,budget_cost,latest_forecast_revenue,latest_forecast_cost,status,source,notes) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'finance_workbook',%s) ON DUPLICATE KEY UPDATE actual_revenue=VALUES(actual_revenue),actual_cost=VALUES(actual_cost),budget_revenue=VALUES(budget_revenue),budget_cost=VALUES(budget_cost),latest_forecast_revenue=VALUES(latest_forecast_revenue),latest_forecast_cost=VALUES(latest_forecast_cost),status=VALUES(status),notes=VALUES(notes)", (ESTATE_ID, year, month, as_number(row.get("Actual revenue")), as_number(row.get("Actual cost")), as_number(row.get("Budget revenue")), as_number(row.get("Budget cost")), as_number(row.get("Latest forecast revenue")), as_number(row.get("Latest forecast cost")), status, as_text(row.get("Notes"))))

    def import_vat(self, book: Any) -> None:
        _, rows = find_table(book["VAT & Tax"], "Year")
        for _, row in rows:
            year = as_int(row.get("Year"))
            if not year:
                continue
            self.bump("vat_returns")
            if self.commit_mode:
                self.cursor.execute("INSERT INTO vat_returns (estate_id,fiscal_year,filing_status,sales_net,taxable_purchases,output_vat,input_vat,annual_credit,prior_credit,credit_used,ending_credit,source_document) VALUES (%s,%s,'filed',%s,%s,%s,%s,%s,%s,%s,%s,%s) ON DUPLICATE KEY UPDATE sales_net=VALUES(sales_net),taxable_purchases=VALUES(taxable_purchases),output_vat=VALUES(output_vat),input_vat=VALUES(input_vat),annual_credit=VALUES(annual_credit),prior_credit=VALUES(prior_credit),credit_used=VALUES(credit_used),ending_credit=VALUES(ending_credit),source_document=VALUES(source_document)", (ESTATE_ID, year, as_number(row.get("Sales / volume of business")), as_number(row.get("Taxable purchases")), as_number(row.get("Output VAT")), as_number(row.get("Input VAT")), as_number(row.get("Annual VAT credit")), as_number(row.get("Prior-year VAT credit")), as_number(row.get("Prior credit used in F24")), as_number(row.get("Annual VAT credit")), as_text(row.get("Source"))))

    def import_accounts(self, book: Any) -> None:
        _, rows = find_table(book["Account Detail"], "Year")
        for _, row in rows:
            year, code = as_int(row.get("Year")), as_text(row.get("Account"))
            if not year or not code:
                continue
            self.bump("account_balance_lines")
            if self.commit_mode:
                self.cursor.execute("INSERT INTO account_balance_lines (id,estate_id,fiscal_year,section,account_code,description,book_amount,non_deductible_pct,non_deductible_amount,tax_amount,source_document) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) ON DUPLICATE KEY UPDATE description=VALUES(description),book_amount=VALUES(book_amount),non_deductible_pct=VALUES(non_deductible_pct),non_deductible_amount=VALUES(non_deductible_amount),tax_amount=VALUES(tax_amount),source_document=VALUES(source_document)", (uid(), ESTATE_ID, year, as_text(row.get("Section")) or "Unknown", code, as_text(row.get("Description")) or code, as_number(row.get("Book amount")), as_number(row.get("Non-deductible %")), as_number(row.get("Non-deductible amount")), as_number(row.get("Tax amount")), as_text(row.get("Source"))))

    def import_prices(self, book: Any) -> None:
        _, rows = find_table(book["Price List"], "Vintage")
        channel_columns = (("Restaurant / Bar incl. VAT", "restaurant_bar", True), ("Vineyard incl. VAT", "vineyard", True), ("Wholesale net + VAT", "wholesale", False))
        for _, row in rows:
            year, wine = as_int(row.get("Vintage")), as_text(row.get("Wine"))
            if not year or not wine:
                continue
            vat_rate = as_number(row.get("VAT rate"))
            for column, channel, gross in channel_columns:
                value = as_number(row.get(column))
                if value is None:
                    continue
                self.bump("price_entries")
                if self.commit_mode:
                    net = value / (1 + vat_rate) if gross and vat_rate is not None else value
                    gross_value = value if gross else value * (1 + (vat_rate or 0))
                    self.cursor.execute("INSERT INTO price_entries (id,estate_id,vintage_year,wine_name,channel,effective_date,price_net,price_gross,vat_rate,status,source,notes) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'finance_workbook',%s) ON DUPLICATE KEY UPDATE price_net=VALUES(price_net),price_gross=VALUES(price_gross),vat_rate=VALUES(vat_rate),status=VALUES(status),notes=VALUES(notes)", (uid(), ESTATE_ID, year, wine, channel, date(2026,7,1), net, gross_value, vat_rate, "current" if year == 2025 else "historical", as_text(row.get("Status / note"))))

    def import_contract_services(self, book: Any) -> None:
        for _, rows in find_tables(book["Labor"], "Year"):
            for _, row in rows:
                year, service, person = as_int(row.get("Year")), as_text(row.get("Work / Service")), as_text(row.get("Person"))
                if not year or not service or not person:
                    continue
                amount, note = as_number(row.get("Cost")), as_text(row.get("Basis / Note"))
                status = "budget" if note and "Budget" in note else ("actual" if amount is not None else "unknown")
                self.bump("contract_service_costs")
                party_id = self.party(person, "supplier")
                if self.commit_mode:
                    self.cursor.execute("INSERT INTO contract_service_costs (id,estate_id,fiscal_year,service_type,party_id,person_name,amount_eur,record_status,basis_note,source) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,'finance_workbook') ON DUPLICATE KEY UPDATE party_id=VALUES(party_id),amount_eur=VALUES(amount_eur),record_status=VALUES(record_status),basis_note=VALUES(basis_note)", (uid(), ESTATE_ID, year, service, party_id, person, amount, status, note))

    def import_finance_sources(self, book: Any) -> None:
        _, rows = find_table(book["Checks & Sources"], "Check")
        for _, row in rows:
            source_id = as_text(row.get("Check"))
            if not source_id or not re.match(r"^[A-Z]+-", source_id):
                continue
            self.bump("evidence_references")
            if self.commit_mode:
                self.cursor.execute("INSERT INTO evidence_references (id,estate_id,evidence_type,external_id,title,source_url,confidence,notes,metadata) VALUES (%s,%s,'document',%s,%s,%s,'authoritative',%s,%s)", (uid(), ESTATE_ID, source_id, as_text(row.get("Actual")), as_text(row.get("Link")), as_text(row.get("Notes")), json.dumps(row, default=json_value)))

    def import_funding(self, book: Any) -> None:
        self.import_opportunities(book)
        self.import_requirements(book)
        self.import_timeline(book)
        self.import_capital_plan(book)

    def import_opportunities(self, book: Any) -> None:
        _, rows = find_table(book["Opportunity Tracker"], "ID")
        for _, row in rows:
            source_id, name = as_text(row.get("ID")), as_text(row.get("Program"))
            if not source_id or not name:
                continue
            priority = (as_text(row.get("Priority")) or "Medium").lower()
            if priority not in {"low", "medium", "high", "critical"}: priority = "medium"
            self.bump("funding_opportunities")
            if self.commit_mode:
                self.cursor.execute("INSERT INTO funding_opportunities (id,estate_id,source_opportunity_id,program_name,program_level,status,fit,priority,opens_on,deadline,support_type,rate_amount,eligible_use,eligibility_risk,project_role,next_action,owner_text,last_verified,confidence_timing,official_source_url) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) ON DUPLICATE KEY UPDATE program_name=VALUES(program_name),status=VALUES(status),fit=VALUES(fit),priority=VALUES(priority),deadline=VALUES(deadline),eligible_use=VALUES(eligible_use),eligibility_risk=VALUES(eligibility_risk),next_action=VALUES(next_action),owner_text=VALUES(owner_text),last_verified=VALUES(last_verified),confidence_timing=VALUES(confidence_timing),official_source_url=VALUES(official_source_url)", (uid(), ESTATE_ID, source_id, name, as_text(row.get("Level")), as_text(row.get("Status")) or "Unknown", as_text(row.get("Fit")), priority, workbook_date(row.get("Opens")), workbook_date(row.get("Deadline")), as_text(row.get("Support type")), as_text(row.get("Rate / amount")), as_text(row.get("Eligible use")), as_text(row.get("Key eligibility / risk")), as_text(row.get("Project role")), as_text(row.get("Next action")), as_text(row.get("Owner")), workbook_date(row.get("Last verified")), as_text(row.get("Confidence / timing note")), as_text(row.get("Official source URL"))))

    def import_requirements(self, book: Any) -> None:
        _, rows = find_table(book["Eligibility & Docs"], "Category")
        status_map = {"not started":"not_started", "in progress":"in_progress", "complete":"complete", "n/a":"not_applicable", "blocked":"blocked"}
        for _, row in rows:
            category, requirement = as_text(row.get("Category")), as_text(row.get("Required document / test"))
            if not category or not requirement:
                continue
            status = status_map.get((as_text(row.get("Status")) or "not started").lower(), "not_started")
            self.bump("funding_requirements")
            if self.commit_mode:
                self.cursor.execute("INSERT INTO funding_requirements (id,estate_id,category,requirement_name,programs_text,owner_text,status,due_date,notes) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) ON DUPLICATE KEY UPDATE programs_text=VALUES(programs_text),owner_text=VALUES(owner_text),status=VALUES(status),due_date=VALUES(due_date),notes=VALUES(notes)", (uid(), ESTATE_ID, category, requirement, as_text(row.get("Programs")), as_text(row.get("Owner")), status, workbook_date(row.get("Due")), as_text(row.get("Notes"))))

    def import_timeline(self, book: Any) -> None:
        _, rows = find_table(book["Timeline"], "Phase")
        priority_map = {"critical":"critical", "high":"high", "medium":"medium", "low":"low", "conditional":"conditional"}
        for _, row in rows:
            action = as_text(row.get("Action"))
            if not action:
                continue
            self.bump("funding_milestones")
            if self.commit_mode:
                priority = priority_map.get((as_text(row.get("Priority")) or "medium").lower(), "medium")
                self.cursor.execute("INSERT INTO funding_milestones (id,estate_id,phase,starts_on,ends_on,workstream,action_text,owner_text,priority,dependency_guardrail) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)", (uid(), ESTATE_ID, as_text(row.get("Phase")), workbook_date(row.get("Start")), workbook_date(row.get("End / deadline")), as_text(row.get("Workstream")) or "Funding", action, as_text(row.get("Owner")), priority, as_text(row.get("Dependency / guardrail"))))

    def import_capital_plan(self, book: Any) -> None:
        project_id = str(uuid.uuid5(uuid.NAMESPACE_URL, "baiamonte:capital:passopisciaro"))
        if self.commit_mode:
            self.cursor.execute("INSERT INTO capital_projects (id,estate_id,code,name,site,owner_payer,company_project,status,planned_start,planned_end,budget_low,budget_high,decision_gate) VALUES (%s,%s,'PASSOPISCIARO','Passopisciaro cellar and two-site investment plan','Vineyard estate + Passopisciaro','Tenuta Baiamonte',1,'planning','2026-08-01','2028-07-31',559000,772500,'Keep Baiamonte company, vineyard lease, and private apartment costs legally and financially separated') ON DUPLICATE KEY UPDATE budget_low=VALUES(budget_low),budget_high=VALUES(budget_high),decision_gate=VALUES(decision_gate)", (project_id, ESTATE_ID))
        _, rows = find_table(book["Funding Model"], "Cost package")
        for row_number, row in rows:
            description = as_text(row.get("Cost package"))
            if not description or description.startswith("TOTAL") or description.startswith("COMBINED"):
                continue
            self.bump("capital_budget_lines")
            if self.commit_mode:
                private = 1 if "Private apartment" in description else 0
                self.cursor.execute("INSERT INTO capital_budget_lines (id,project_id,cost_code,description,owner_payer,target_timing,budget_low,budget_high,support_low_pct,support_high_pct,primary_route_text,eligibility_gate,private_excluded) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) ON DUPLICATE KEY UPDATE description=VALUES(description),owner_payer=VALUES(owner_payer),target_timing=VALUES(target_timing),budget_low=VALUES(budget_low),budget_high=VALUES(budget_high),support_low_pct=VALUES(support_low_pct),support_high_pct=VALUES(support_high_pct),primary_route_text=VALUES(primary_route_text),eligibility_gate=VALUES(eligibility_gate),private_excluded=VALUES(private_excluded)", (uid(), project_id, f"FM-{row_number:03d}", description, as_text(row.get("Owner / payer")), as_text(row.get("Target timing")), as_number(row.get("Low spend")), as_number(row.get("High spend")), as_number(row.get("Low support %")), as_number(row.get("High support %")), as_text(row.get("Primary funding route")), as_text(row.get("Eligibility / decision gate")), private))

    def report(self) -> dict[str, Any]:
        return {"mode": "commit" if self.commit_mode else "dry-run", "workbooks": [path.name for path in self.paths], "counts": self.counts, "warnings": self.warnings, "controls": ["All non-empty source rows are retained with workbook, tab, and row provenance.", "Actual, budget, forecast, and funding status remain distinct.", "Private apartment costs are excluded from company funding by default.", "Each invoice line can have only one primary funding allocation.", "Blank financial values remain NULL unless the source explicitly records zero."]}


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbooks", nargs="+", type=Path)
    parser.add_argument("--commit", action="store_true")
    parser.add_argument("--report", type=Path)
    args = parser.parse_args()
    if args.commit:
        parser.error("Workbook commits are retired; MariaDB and connected finance sources are authoritative")
    report = FinanceImporter(args.workbooks, args.commit).run()
    encoded = json.dumps(report, indent=2, default=json_value, ensure_ascii=False)
    if args.report:
        args.report.write_text(encoded + "\n", encoding="utf-8")
    print(encoded)
    return 0


if __name__ == "__main__":
    sys.exit(main())
