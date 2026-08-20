"""Import Baiamonte-only legacy work, compensation, payment, and cost history.

Dry-run is the default.  Societa La Nave sheets and mixed TNB/La Nave rows are
intentionally excluded.  Source totals and cash transfers are not expenses.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import uuid
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import openpyxl
import pymysql

from import_workbook import ESTATE_ID, as_number, as_text, json_value


WORK_FILE_ID = "1hBQy5GUNw1yoSOod944tr42tm1WruLHvE8Wpdok2WnM"
COST_FILE_ID = "1shs_5UC2w3aLp0hMnb2yIITdjNKkOp4a-VwrLVH9xs4"
SUMMARY_WORDS = ("totale", "total from", "work total", "pagato societa", "pagato federico", "closed")
MONTHS = {"dicembre": 12, "gennaio": 1, "febbraio": 2, "marzo": 3, "aprile": 4}


def text(value: Any) -> str:
    return (as_text(value) or "").strip()


def norm(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", text(value).casefold()).strip()


def money(value: Any) -> float | None:
    if isinstance(value, str):
        raw = value.strip().replace("€", "").replace(" ", "")
        if "," in raw and "." not in raw:
            raw = raw.replace(",", ".")
        elif "," in raw and "." in raw:
            raw = raw.replace(".", "").replace(",", ".")
        try:
            return round(abs(float(raw)), 2)
        except ValueError:
            return None
    number = as_number(value)
    return None if number is None else round(abs(float(number)), 2)


def parse_date(value: Any, default_year: int | None = None) -> tuple[date | None, str]:
    if isinstance(value, datetime):
        return value.date(), "day"
    if isinstance(value, date):
        return value, "day"
    if isinstance(value, (int, float)):
        if 1000 <= float(value) <= 2200:
            return date(int(value), 1, 1), "year"
        if 1 < float(value) < 100000:
            return date(1899, 12, 30) + timedelta(days=int(value)), "day"
    raw = text(value)
    if not raw:
        return (date(default_year, 1, 1), "year") if default_year else (None, "unknown")
    first = re.split(r"\s*/\s*(?=\d{1,2}-\d{1,2}-\d{2,4})", raw)[0]
    for fmt in ("%d-%m-%Y", "%d-%m-%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(first, fmt).date(), "day"
        except ValueError:
            pass
    slash = re.fullmatch(r"(\d{1,2})/(\d{1,2})/(\d{2}|\d{4})", first)
    if slash:
        left, middle, year_value = (int(part) for part in slash.groups())
        year = year_value + 2000 if year_value < 100 else year_value
        # These are Italian workbooks and are day-first unless that would be
        # impossible. A value such as 06/23/25 is safely month-first.
        day, month = (middle, left) if middle > 12 else (left, middle)
        try:
            return date(year, month, day), "day"
        except ValueError:
            pass
    month_year = re.fullmatch(r"(\d{1,2})/(20\d{2})", first)
    if month_year:
        month, year = (int(part) for part in month_year.groups())
        try:
            return date(year, month, 1), "month"
        except ValueError:
            pass
    if re.fullmatch(r"20\d{2}", raw):
        return date(int(raw), 1, 1), "year"
    return (date(default_year, 1, 1), "year") if default_year else (None, "unknown")


def month_sheet_date(value: Any, year: int, month: int) -> tuple[date, str]:
    parsed, precision = parse_date(value)
    if not parsed:
        return date(year, month, 1), "month"
    if parsed.year == year and parsed.month == month:
        return parsed, precision
    # These source tabs were entered as day/month but parsed by Sheets as
    # month/day.  The tab month is authoritative; the parsed month is the day.
    day = parsed.month if 1 <= parsed.month <= 31 else 1
    return date(year, month, min(day, 28 if month == 2 else 30 if month in {4, 6, 9, 11} else 31)), "day"


def actor(description: str) -> str | None:
    clean = norm(description)
    for needle, name in (("sonia", "Sonia"), ("federico", "Federico"), ("giancarlo", "Giancarlo"), ("filippo", "Filippo"), ("emanuele", "Emanuele"), ("ture", "Ture"), ("luca", "Luca")):
        if re.search(rf"\b{needle}\b", clean):
            return name
    return None


def classify(description: str) -> str:
    value = norm(description)
    groups = (
        ("olive_oil", ("olive", "olio", "oleficio", "molitura")),
        ("wine_cellar", ("vino", "bottigl", "etichett", "pallet", "cella frigo", "wine")),
        ("fuel_water", ("carburante", "gasolio", "benzina", "diesel", "acqua", "gas ")),
        ("vineyard_inputs", ("vigna", "uva", "piante", "pali", "concime", "trattament", "potatura", "motozappa", "legato")),
        ("labor", ("lavoro", "manodopera", "giorni", "operai", "giancarlo", "filippo", "emanuele", "ture", "sonia", "federico")),
        ("capital_improvement", ("impianto", "escavatore", "cisterna", "fontana", "ingegnere", "struttura", "cancello", "gate", "elettrico", "edile")),
        ("transport_shipping", ("trasporto", "spedizion", "shipping", "dhl")),
    )
    return next((name for name, words in groups if any(word in value for word in words)), "materials_other")


@dataclass
class Record:
    source_file_id: str
    source_file_name: str
    source_sheet: str
    source_row_number: int
    description: str
    amount_eur: float
    labor_hours: float | None = None
    source_item_key: str = "row"
    record_kind: str = "expense"
    record_date: date | None = None
    record_year: int | None = None
    period_start_year: int | None = None
    period_end_year: int | None = None
    date_precision: str = "unknown"
    classification: str | None = None
    actor_name: str | None = None
    payment_method: str | None = None
    payment_status: str | None = None
    included_in_totals: bool = True
    exclusion_reason: str | None = None
    raw_values: list[Any] = field(default_factory=list)
    id: str = field(init=False)
    source_row_hash: str = field(init=False)
    duplicate_of: str | None = None

    def __post_init__(self) -> None:
        self.classification = self.classification or classify(self.description)
        self.actor_name = self.actor_name or actor(self.description)
        self.record_year = self.record_year or (self.record_date.year if self.record_date else None)
        if self.labor_hours is None:
            hours = re.search(r"\b(\d+(?:[.,]\d+)?)\s*(?:ore|hours?)\b", self.description, re.IGNORECASE)
            self.labor_hours = float(hours.group(1).replace(",", ".")) if hours else None
        key = f"{self.source_file_id}:{self.source_sheet}:{self.source_row_number}:{self.source_item_key}"
        self.id = str(uuid.uuid5(uuid.NAMESPACE_URL, f"baiamonte:legacy-cost:{key}"))
        payload = json.dumps(self.raw_values, default=json_value, ensure_ascii=False)
        self.source_row_hash = hashlib.sha256(payload.encode()).hexdigest()


def is_summary(description: str) -> bool:
    value = norm(description)
    return not value or any(word in value for word in SUMMARY_WORDS) or value in {"work paid", "work ongoing", "pre paid total", "pagato"}


def parse_work_book(path: Path) -> list[Record]:
    book = openpyxl.load_workbook(path, data_only=True, read_only=True)
    records: list[Record] = []
    for sheet in book.worksheets:
        sheet_name = sheet.title
        lower = sheet_name.casefold()
        year_match = re.search(r"20\d{2}", sheet_name)
        sheet_year = int(year_match.group()) if year_match else None
        if "societ" in lower and "nave" in lower:
            continue
        if "lavori eseguiti" in lower:
            for row_number, values in enumerate(sheet.iter_rows(values_only=True), 1):
                values = list(values)
                description, amount = text(values[1] if len(values) > 1 else None), money(values[2] if len(values) > 2 else None)
                if amount is None or amount <= 0 or is_summary(description):
                    continue
                record_date, precision = parse_date(values[3] if len(values) > 3 else None, sheet_year)
                records.append(Record(WORK_FILE_ID, path.name, sheet_name, row_number, description, amount, record_date=record_date, date_precision=precision, payment_status=text(values[4] if len(values) > 4 else None) or None, raw_values=values))
        elif "payments" in lower:
            in_payments = False
            for row_number, values in enumerate(sheet.iter_rows(values_only=True), 1):
                values = list(values)
                description = text(values[1] if len(values) > 1 else None)
                amount = money(values[2] if len(values) > 2 else None)
                if norm(description) == "payments":
                    in_payments = True
                    continue
                if amount is None or amount <= 0 or is_summary(description) or "balance forward" in norm(description):
                    continue
                if norm(description).startswith("compenso"):
                    person = actor(description)
                    if not person:
                        continue
                    years = [int(y) for y in re.findall(r"20\d{2}", description)]
                    start = min(years) if years else sheet_year
                    end = max(years) if years else sheet_year
                    single_year = start if start == end else None
                    records.append(Record(WORK_FILE_ID, path.name, sheet_name, row_number, description, amount, record_kind="compensation", record_year=single_year, period_start_year=start, period_end_year=end, date_precision="year" if single_year else "period", actor_name=person, raw_values=values))
                elif in_payments:
                    person = actor(description)
                    record_date, precision = parse_date(values[3] if len(values) > 3 else None, sheet_year)
                    records.append(Record(WORK_FILE_ID, path.name, sheet_name, row_number, description, amount, record_kind="payment", record_date=record_date, date_precision=precision, actor_name=person, payment_method=text(values[4] if len(values) > 4 else None) or None, included_in_totals=False, exclusion_reason="Settlement/payment retained separately from expense totals", raw_values=values))
    return records


def parse_cost_book(path: Path) -> list[Record]:
    book = openpyxl.load_workbook(path, data_only=True, read_only=True)
    records: list[Record] = []
    for sheet in book.worksheets:
        sheet_name = sheet.title
        lower = sheet_name.strip().casefold()
        if "la nave" in lower:
            continue
        if lower == "spese sistemazione baia monte":
            for row_number, values in enumerate(sheet.iter_rows(values_only=True), 1):
                values = list(values)
                description, amount = text(values[0] if values else None), money(values[1] if len(values) > 1 else None)
                if row_number == 1 or amount is None or amount <= 0 or is_summary(description):
                    continue
                record_date, precision = parse_date(values[2] if len(values) > 2 else None)
                transfer = "prelievo per ripristino cassa" in norm(description)
                records.append(Record(COST_FILE_ID, path.name, sheet_name, row_number, description, amount, record_kind="transfer" if transfer else "expense", record_date=record_date, date_precision=precision, payment_method=text(values[3] if len(values) > 3 else None) or None, included_in_totals=not transfer, exclusion_reason="Cash replenishment is a transfer, not an expense" if transfer else None, raw_values=values))
        elif lower == "ture-2023":
            values = list(next(sheet.iter_rows(min_row=5, max_row=5, values_only=True)))
            for column, amount_value in enumerate(values[2:14], 1):
                amount = money(amount_value)
                if not amount:
                    continue
                records.append(Record(COST_FILE_ID, path.name, sheet_name, 5, f"Ture bill {date(2023, column, 1).strftime('%B %Y')}", amount, source_item_key=f"month-{column:02d}", record_kind="compensation", record_date=date(2023, column, 1), date_precision="month", actor_name="Ture", raw_values=[column, amount_value]))
        elif lower == "lavori baia monte":
            for row_number, values in enumerate(sheet.iter_rows(values_only=True), 1):
                values = list(values)
                description, amount = text(values[1] if len(values) > 1 else None), money(values[2] if len(values) > 2 else None)
                if amount is None or amount <= 0 or is_summary(description):
                    continue
                record_date, precision = parse_date(values[0] if values else None, 2024)
                records.append(Record(COST_FILE_ID, path.name, sheet_name, row_number, description, amount, record_date=record_date, date_precision=precision, payment_status=text(values[3] if len(values) > 3 else None) or None, raw_values=values))
        else:
            month_name = next((name for name in MONTHS if lower.startswith(name)), None)
            year_match = re.search(r"20\d{2}", lower)
            if not month_name or not year_match:
                continue
            year, month = int(year_match.group()), MONTHS[month_name]
            for row_number, values in enumerate(sheet.iter_rows(values_only=True), 1):
                values = list(values)
                entity = norm(values[3] if len(values) > 3 else None)
                # Only rows owned solely by Tenuta/Baia Monte. Mixed rows are
                # excluded because the workbook does not provide an allocation.
                if entity != "tnb":
                    continue
                description, amount = text(values[1] if len(values) > 1 else None), money(values[2] if len(values) > 2 else None)
                if amount is None or amount <= 0 or is_summary(description):
                    continue
                record_date, precision = month_sheet_date(values[0] if values else None, year, month)
                records.append(Record(COST_FILE_ID, path.name, sheet_name, row_number, description, amount, record_date=record_date, date_precision=precision, payment_method=text(values[4] if len(values) > 4 else None) or None, raw_values=values))
    return records


def reconcile(records: list[Record]) -> None:
    exact: dict[tuple[Any, ...], Record] = {}
    for record in records:
        if not record.included_in_totals:
            continue
        key = (record.record_date, record.amount_eur, norm(record.description))
        if key in exact:
            record.included_in_totals = False
            record.duplicate_of = exact[key].id
            record.exclusion_reason = "Exact duplicate of another imported Baiamonte source row"
        else:
            exact[key] = record
    details = [r for r in records if r.included_in_totals and r.actor_name == "Ture" and r.source_sheet.strip().casefold() == "spese sistemazione baia monte"]
    month_words = {1: "gennaio", 2: "febbraio", 3: "marzo", 4: "aprile", 5: "maggio", 6: "giugno", 7: "luglio", 8: "agosto", 9: "settembre", 10: "ottobre", 11: "novembre", 12: "dicembre"}
    for summary in records:
        if not summary.included_in_totals or summary.source_sheet.casefold() != "ture-2023" or not summary.record_date:
            continue
        match = next((r for r in details if r.amount_eur == summary.amount_eur and ((r.record_date and r.record_date.year == summary.record_date.year and r.record_date.month == summary.record_date.month) or month_words[summary.record_date.month] in norm(r.description))), None)
        if match:
            summary.included_in_totals = False
            summary.duplicate_of = match.id
            summary.exclusion_reason = "Monthly Ture summary duplicates a dated Baiamonte expense"


def report(records: list[Record], commit: bool) -> dict[str, Any]:
    by_year: dict[str, dict[str, float | int]] = {}
    for record in records:
        if not record.record_year:
            continue
        item = by_year.setdefault(str(record.record_year), {"records": 0, "expenses_eur": 0.0, "payments_eur": 0.0})
        item["records"] += 1
        if record.included_in_totals:
            item["expenses_eur"] = round(float(item["expenses_eur"]) + record.amount_eur, 2)
        if record.record_kind == "payment":
            item["payments_eur"] = round(float(item["payments_eur"]) + record.amount_eur, 2)
    return {"mode": "commit" if commit else "dry-run", "counts": {"selected_baiamonte_rows": len(records), "included_expense_rows": sum(r.included_in_totals for r in records), "payments": sum(r.record_kind == "payment" for r in records), "duplicates_excluded": sum(bool(r.duplicate_of) for r in records), "transfers_excluded": sum(r.record_kind == "transfer" for r in records)}, "by_year": by_year, "warnings": ["Societa La Nave sheets and mixed TNB/La Nave rows were excluded.", "Payments are retained separately and do not increase expense totals.", "The Federico 2023/2024 compensation remains an unsplit period record because the source does not allocate it by year."]}


def save(records: list[Record]) -> None:
    connection = pymysql.connect(host=os.getenv("DB_HOST", "core-mariadb"), port=int(os.getenv("DB_PORT", "3306")), user=os.getenv("DB_USER", "baiamonte"), password=os.getenv("DB_PASSWORD", ""), database=os.getenv("DB_NAME", "baiamonte_vineyard"), charset="utf8mb4", autocommit=False)
    try:
        with connection.cursor() as cursor:
            for record in records:
                cursor.execute(
                    "INSERT INTO historical_cost_records (id,estate_id,source_file_id,source_file_name,source_sheet,source_row_number,source_item_key,source_row_hash,record_date,record_year,period_start_year,period_end_year,date_precision,record_kind,classification,actor_name,description,amount_eur,labor_hours,payment_method,payment_status,included_in_totals,duplicate_of,exclusion_reason,raw_values) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) ON DUPLICATE KEY UPDATE source_file_name=VALUES(source_file_name),source_row_hash=VALUES(source_row_hash),record_date=VALUES(record_date),record_year=VALUES(record_year),period_start_year=VALUES(period_start_year),period_end_year=VALUES(period_end_year),date_precision=VALUES(date_precision),record_kind=VALUES(record_kind),classification=VALUES(classification),actor_name=VALUES(actor_name),description=VALUES(description),amount_eur=VALUES(amount_eur),labor_hours=VALUES(labor_hours),payment_method=VALUES(payment_method),payment_status=VALUES(payment_status),included_in_totals=VALUES(included_in_totals),duplicate_of=VALUES(duplicate_of),exclusion_reason=VALUES(exclusion_reason),raw_values=VALUES(raw_values)",
                    (record.id, ESTATE_ID, record.source_file_id, record.source_file_name, record.source_sheet, record.source_row_number, record.source_item_key, record.source_row_hash, record.record_date, record.record_year, record.period_start_year, record.period_end_year, record.date_precision, record.record_kind, record.classification, record.actor_name, record.description, record.amount_eur, record.labor_hours, record.payment_method, record.payment_status, record.included_in_totals, record.duplicate_of, record.exclusion_reason, json.dumps(record.raw_values, default=json_value, ensure_ascii=False)),
                )
        connection.commit()
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--work-history", type=Path)
    parser.add_argument("--costs-history", type=Path)
    parser.add_argument("--commit", action="store_true")
    parser.add_argument("--report", type=Path)
    args = parser.parse_args()
    parser.error("Workbook commits are retired and workbook access is retired; MariaDB is authoritative")


if __name__ == "__main__":
    sys.exit(main())
